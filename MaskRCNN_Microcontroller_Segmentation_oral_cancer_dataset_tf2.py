import os
import sys
import random
import math
import imgaug
import json
import xlsxwriter
import re
import time
import openpyxl
import h5py
import numpy as np
import cv2
import glob
import tensorflow as tf
import matplotlib
import matplotlib.pyplot as plt
import xml.etree.ElementTree as ET
import argparse

import matplotlib
import matplotlib.pyplot as plt

# Root directory of the project
ROOT_DIR = os.path.abspath("./")

#os.environ["CUDA_VISIBLE_DEVICES"] = '1'

# Import Mask RCNN
sys.path.append(ROOT_DIR)  # To find local version of the library

from mrcnn.config import Config
from mrcnn import utils

from openpyxl import Workbook
from tqdm import tqdm
#from imantics import Mask
from xml.dom import minidom
#import mrcnn.model as modellib
from mrcnn import visualize
# from mrcnn.model import log
from openpyxl import Workbook

# %matplotlib inline 

# Directory to save logs and trained model
MODEL_DIR = os.path.join(ROOT_DIR, "logs")

# Local path to trained weights file
COCO_MODEL_PATH = os.path.join(ROOT_DIR, "mask_rcnn_coco.h5")
# Download COCO trained weights from Releases if needed
if not os.path.exists(COCO_MODEL_PATH):
    utils.download_trained_weights(COCO_MODEL_PATH)

def opt():
    # Parse command line arguments
    parser = argparse.ArgumentParser(
        description='Train Mask R-CNN to detect carplates.')
    parser.add_argument('--mask_mode', default='MaskRCNN',
                        metavar="<command>",
                        help="'MaskRCNN' or 'Cascade_MaskRCNN'")
    parser.add_argument('--training_or_inference_mode', default='training',
                        metavar="<training,inference>",
                        help='Directory of the Carplate dataset')
    #parser.add_argument('--model_path', required=True,
    #                    metavar="<model file name>",
    #                    help='Directory of the log dataset')
    parser.add_argument('--model_path', default='',
                        metavar="<model file name>",
                        help='Directory of the log dataset')
    parser.add_argument('--backbone', default='resnet50',
                        metavar="<model file name>",
                        help='Directory of the log dataset')
    parser.add_argument('--device', default='',
                        metavar="<model file name>",
                        help='Directory of the log dataset')
    args = parser.parse_args()
    
    return args
    
class ShapesConfig(Config):
    """Configuration for training on the toy shapes dataset.
    Derives from the base Config class and overrides values specific
    to the toy shapes dataset.
    """
    NAME = "microcontroller_detection" # Override in sub-classes
    
    category_name = ['G', 'Y', 'R']
    
    colors = [(0, 1, 0), (1, 1, 0), (1, 0, 0)]
    colorlist = ['Gray', 'Green', 'Yellow', 'Red']
    
    augmentation_types = ['vertical', 'horizontal', 'horizontal_and_vertical']

    # NUMBER OF GPUs to use. When using only a CPU, this needs to be set to 1.
    GPU_COUNT = 1
    
    # Use Multiprocessing in MaskRCNN.train()
    USE_MULTIPROCESSING = False
    
    N_STEPS = 5

    # Number of images to train with on each GPU. A 12GB GPU can typically
    # handle 2 images of 1024x1024px.
    # Adjust based on your GPU memory and image sizes. Use the highest
    # number that your GPU can handle for best performance.
    IMAGES_PER_GPU = 1
    
    AMT_TRAIN = 3000
    AMT_VAL = 200
    AMT_SMALL_VAL = 200

    # Number of training steps per epoch
    # This doesn't need to match the size of the training set. Tensorboard
    # updates are saved at the end of each epoch, so setting this to a
    # smaller number means getting more frequent TensorBoard updates.
    # Validation stats are also calculated at each epoch end and they
    # might take a while, so don't set this too small to avoid spending
    # a lot of time on validation stats.
    STEPS_PER_EPOCH = 1000

    # Number of validation steps to run at the end of every training epoch.
    # A bigger number improves accuracy of validation stats, but slows
    # down the training.
    VALIDATION_STEPS = 50

    # Backbone network architecture
    # Supported values are: resnet50, resnet101.
    # You can also provide a callable that should have the signature
    # of model.resnet_graph. If you do so, you need to supply a callable
    # to COMPUTE_BACKBONE_SHAPE as well
    BACKBONE = "resnet101"
    #backbone = "resnet"

    # Only useful if you supply a callable to BACKBONE. Should compute
    # the shape of each layer of the FPN Pyramid.
    # See model.compute_backbone_shapes
    COMPUTE_BACKBONE_SHAPE = None
    
    DETECTION_HEAD = "original"
    MASK_HEAD = "original"

    # The strides of each layer of the FPN Pyramid. These values
    # are based on a Resnet101 backbone.
    BACKBONE_STRIDES = [4, 8, 16, 32, 64]

    # Size of the fully-connected layers in the classification graph
    FPN_CLASSIF_FC_LAYERS_SIZE = 1024

    # Size of the top-down layers used to build the feature pyramid
    TOP_DOWN_PYRAMID_SIZE = 256

    # Number of classification classes (including background)
    NUM_CLASSES = 4  # Override in sub-classes
    NUM_KEYPOINTS = 17
    NUM_CLASSES_UNET = 3

    # Length of square anchor side in pixels
    RPN_ANCHOR_SCALES = (32, 64, 128, 256, 512)

    # Ratios of anchors at each cell (width/height)
    # A value of 1 represents a square anchor, and 0.5 is a wide anchor
    RPN_ANCHOR_RATIOS = [0.5, 1, 2]

    # Anchor stride
    # If 1 then anchors are created for each cell in the backbone feature map.
    # If 2, then anchors are created for every other cell, and so on.
    RPN_ANCHOR_STRIDE = 1
    
    IMAGE_PADDING = True

    # Non-max suppression threshold to filter RPN proposals.
    # You can increase this during training to generate more propsals.
    RPN_NMS_THRESHOLD = 0.7

    # How many anchors per image to use for RPN training
    RPN_TRAIN_ANCHORS_PER_IMAGE = 256
    
    # ROIs kept after tf.nn.top_k and before non-maximum suppression
    PRE_NMS_LIMIT = 6000

    # ROIs kept after non-maximum suppression (training and inference)
    POST_NMS_ROIS_TRAINING = 2000
    POST_NMS_ROIS_INFERENCE = 1000
    
    # Image mean (RGB)
    #MEAN_PIXEL = np.array([123.7, 116.8, 103.9])
    MEAN_PIXEL_LOL = np.array([90, 91, 70, 141])
    VARIANCE_LOL = np.array([3587, 3146, 2022, 6594])

    # If enabled, resizes instance masks to a smaller size to reduce
    # memory load. Recommended when using high-resolution images.
    USE_MINI_MASK = True
    
    #if BACKBONE == 'resnet50' or BACKBONE == 'resnet101':
    MINI_MASK_SHAPE = (28, 28)  # (height, width) of the mini-mask
    #elif BACKBONE == 'densenet' or BACKBONE == 'mobilenetv1' or BACKBONE == 'mobilenetv2':
    #    MINI_MASK_SHAPE = (56, 56)

    # Input image resizing
    # Generally, use the "square" resizing mode for training and predicting
    # and it should work well in most cases. In this mode, images are scaled
    # up such that the small side is = IMAGE_MIN_DIM, but ensuring that the
    # scaling doesn't make the long side > IMAGE_MAX_DIM. Then the image is
    # padded with zeros to make it a square so multiple images can be put
    # in one batch.
    # Available resizing modes:
    # none:   No resizing or padding. Return the image unchanged.
    # square: Resize and pad with zeros to get a square image
    #         of size [max_dim, max_dim].
    # pad64:  Pads width and height with zeros to make them multiples of 64.
    #         If IMAGE_MIN_DIM or IMAGE_MIN_SCALE are not None, then it scales
    #         up before padding. IMAGE_MAX_DIM is ignored in this mode.
    #         The multiple of 64 is needed to ensure smooth scaling of feature
    #         maps up and down the 6 levels of the FPN pyramid (2**6=64).
    # crop:   Picks random crops from the image. First, scales the image based
    #         on IMAGE_MIN_DIM and IMAGE_MIN_SCALE, then picks a random crop of
    #         size IMAGE_MIN_DIM x IMAGE_MIN_DIM. Can be used in training only.
    #         IMAGE_MAX_DIM is not used in this mode.
    IMAGE_RESIZE_MODE = "square" # "rectangular", "square", "pad64", "crop"
    IMAGE_MIN_DIM = 856
    IMAGE_MAX_DIM = 1024
    
    # Minimum scaling ratio. Checked after MIN_IMAGE_DIM and can force further
    # up scaling. For example, if set to 2 then images are scaled up to double
    # the width and height, or more, even if MIN_IMAGE_DIM doesn't require it.
    # However, in 'square' mode, it can be overruled by IMAGE_MAX_DIM.
    IMAGE_MIN_SCALE = 0
    # Number of color channels per image. RGB = 3, grayscale = 1, RGB-D = 4
    # Changing this requires other changes in the code. See the WIKI for more
    # details: https://github.com/matterport/Mask_RCNN/wiki
    IMAGE_CHANNEL_COUNT = 3
    
    N_CHANNELS = IMAGE_CHANNEL_COUNT

    # Image mean (RGB)
    MEAN_PIXEL = np.array([123.7, 116.8, 103.9])

    # Number of ROIs per image to feed to classifier/mask heads
    # The Mask RCNN paper uses 512 but often the RPN doesn't generate
    # enough positive proposals to fill this and keep a positive:negative
    # ratio of 1:3. You can increase the number of proposals by adjusting
    # the RPN NMS threshold.
    TRAIN_ROIS_PER_IMAGE = 200

    # Percent of positive ROIs used to train classifier/mask heads
    ROI_POSITIVE_RATIO = 0.5

    # Pooled ROIs
    POOL_SIZE = 7
    MASK_POOL_SIZE = 14

    # Shape of output mask
    # To change this you also need to change the neural network mask branch
    MASK_SHAPE = [28, 28]

    # Maximum number of ground truth instances to use in one image
    MAX_GT_INSTANCES = 100

    # Bounding box refinement standard deviation for RPN and final detections.
    RPN_BBOX_STD_DEV = np.array([0.1, 0.1, 0.2, 0.2])
    BBOX_STD_DEV = np.array([0.1, 0.1, 0.2, 0.2])

    # Max number of final detections
    DETECTION_MAX_INSTANCES = 100

    # Minimum probability value to accept a detected instance
    # ROIs below this threshold are skipped
    DETECTION_MIN_CONFIDENCE = 0.7

    # Non-maximum suppression threshold for detection
    DETECTION_NMS_THRESHOLD = 0.3

    # Learning rate and momentum
    # The Mask RCNN paper uses lr=0.02, but on TensorFlow it causes
    # weights to explode. Likely due to differences in optimizer
    # implementation.
    LEARNING_RATE = 0.008
    LEARNING_MOMENTUM = 0.9

    # Weight decay regularization
    WEIGHT_DECAY = 0.0001

    # Loss weights for more precise optimization.
    # Can be used for R-CNN training setup.
    #if BACKBONE == 'resnet50' or BACKBONE == 'resnet101' or BACKBONE == 'mobilenetv1' or BACKBONE == 'mobilenetv2':
    LOSS_WEIGHTS = {
        "rpn_class_loss": 1.,
        "rpn_bbox_loss": 1.,
        "mrcnn_class_loss": 1.,
        "mrcnn_bbox_loss": 1.,
        "mrcnn_mask_loss": 1.
    }
    #elif BACKBONE == 'densenet':
    #    LOSS_WEIGHTS = {
    #    "rpn_class_loss": 1.,
    #    "rpn_bbox_loss": 1.,
    #    "mrcnn_class_loss": 1.,
    #    "mrcnn_bbox_loss": 1.,
    #    "keypoint_mrcnn_mask_loss": 1.,
    #    "mrcnn_mask_loss": 1.
    #    }

    # Use RPN ROIs or externally generated ROIs for training
    # Keep this True for most situations. Set to False if you want to train
    # the head branches on ROI generated by code rather than the ROIs from
    # the RPN. For example, to debug the classifier head without having to
    # train the RPN.
    USE_RPN_ROIS = True

    # Train or freeze batch normalization layers
    #     None: Train BN layers. This is the normal mode
    #     False: Freeze BN layers. Good when using a small batch size
    #     True: (don't use). Set layer in training mode even when predicting
    TRAIN_BN = False  # Defaulting to False since batch size is often small
    
    KEYPOINT_MASK_SHAPE = [56, 56]
    KEYPOINT_MASK_POOL_SIZE = 17
    
    WEIGHT_LOSS = True

    # Gradient norm clipping
    GRADIENT_CLIP_NORM = 5.0

    def __init__(self):
        """Set values of computed attributes."""
        # Effective batch size
        self.BATCH_SIZE = self.IMAGES_PER_GPU * self.GPU_COUNT

        # Input image size
        if self.IMAGE_RESIZE_MODE == "crop":
            self.IMAGE_SHAPE = np.array([self.IMAGE_MIN_DIM, self.IMAGE_MIN_DIM,
                self.IMAGE_CHANNEL_COUNT])
        else:
            self.IMAGE_SHAPE = np.array([self.IMAGE_MAX_DIM, self.IMAGE_MAX_DIM,
                self.IMAGE_CHANNEL_COUNT])
        
        # Compute backbone size from input image size
        self.BACKBONE_SHAPES = np.array(
            [[int(math.ceil(self.IMAGE_SHAPE[0] / stride)),
              int(math.ceil(self.IMAGE_SHAPE[1] / stride))]
             for stride in self.BACKBONE_STRIDES])

        # Image meta data length
        # See compose_image_meta() for details
        self.IMAGE_META_SIZE = 1 + 3 + 3 + 4 + 1 + self.NUM_CLASSES

    def display(self):
        """Display Configuration values."""
        print("\nConfigurations:")
        for a in dir(self):
            if not a.startswith("__") and not callable(getattr(self, a)):
                print("{:30} {}".format(a, getattr(self, a)))
        print("\n")

def get_ax(rows=1, cols=1, size=8):
    """Return a Matplotlib Axes array to be used in
    all visualizations in the notebook. Provide a
    central point to control graph sizes.
    
    Change the default size attribute to control the size
    of rendered images
    """
    _, ax = plt.subplots(rows, cols, figsize=(size*cols, size*rows))
    return ax

'''
class ShapesDataset(utils.Dataset):
    """Generates the shapes synthetic dataset. The dataset consists of simple
    shapes (triangles, squares, circles) placed randomly on a blank surface.
    The images are generated on the fly. No file access required.
    """

    def load_shapes(self, count, height, width):
        """Generate the requested number of synthetic images.
        count: number of images to generate.
        height, width: the size of the generated images.
        """
        # Add classes
        self.add_class("shapes", 1, "square")
        self.add_class("shapes", 2, "circle")
        self.add_class("shapes", 3, "triangle")

        # Add images
        # Generate random specifications of images (i.e. color and
        # list of shapes sizes and locations). This is more compact than
        # actual images. Images are generated on the fly in load_image().
        for i in range(count):
            bg_color, shapes = self.random_image(height, width)
            self.add_image("shapes", image_id=i, path=None,
                           width=width, height=height,
                           bg_color=bg_color, shapes=shapes)

    def load_image(self, image_id):
        """Generate an image from the specs of the given image ID.
        Typically this function loads the image from a file, but
        in this case it generates the image on the fly from the
        specs in image_info.
        """
        info = self.image_info[image_id]
        bg_color = np.array(info['bg_color']).reshape([1, 1, 3])
        image = np.ones([info['height'], info['width'], 3], dtype=np.uint8)
        image = image * bg_color.astype(np.uint8)
        for shape, color, dims in info['shapes']:
            image = self.draw_shape(image, shape, dims, color)
        return image

    def image_reference(self, image_id):
        """Return the shapes data of the image."""
        info = self.image_info[image_id]
        if info["source"] == "shapes":
            return info["shapes"]
        else:
            super(self.__class__).image_reference(self, image_id)

    def load_mask(self, image_id):
        """Generate instance masks for shapes of the given image ID.
        """
        info = self.image_info[image_id]
        shapes = info['shapes']
        count = len(shapes)
        mask = np.zeros([info['height'], info['width'], count], dtype=np.uint8)
        for i, (shape, _, dims) in enumerate(info['shapes']):
            mask[:, :, i:i+1] = self.draw_shape(mask[:, :, i:i+1].copy(),
                                                shape, dims, 1)
        # Handle occlusions
        occlusion = np.logical_not(mask[:, :, -1]).astype(np.uint8)
        for i in range(count-2, -1, -1):
            mask[:, :, i] = mask[:, :, i] * occlusion
            occlusion = np.logical_and(occlusion, np.logical_not(mask[:, :, i]))
        # Map class names to class IDs.
        class_ids = np.array([self.class_names.index(s[0]) for s in shapes])
        return mask.astype(np.bool), class_ids.astype(np.int32)

    def draw_shape(self, image, shape, dims, color):
        """Draws a shape from the given specs."""
        # Get the center x, y and the size s
        x, y, s = dims
        if shape == 'square':
            cv2.rectangle(image, (x-s, y-s), (x+s, y+s), color, -1)
        elif shape == "circle":
            cv2.circle(image, (x, y), s, color, -1)
        elif shape == "triangle":
            points = np.array([[(x, y-s),
                                (x-s/math.sin(math.radians(60)), y+s),
                                (x+s/math.sin(math.radians(60)), y+s),
                                ]], dtype=np.int32)
            cv2.fillPoly(image, points, color)
        return image

    def random_shape(self, height, width):
        """Generates specifications of a random shape that lies within
        the given height and width boundaries.
        Returns a tuple of three valus:
        * The shape name (square, circle, ...)
        * Shape color: a tuple of 3 values, RGB.
        * Shape dimensions: A tuple of values that define the shape size
                            and location. Differs per shape type.
        """
        # Shape
        shape = random.choice(["square", "circle", "triangle"])
        # Color
        color = tuple([random.randint(0, 255) for _ in range(3)])
        # Center x, y
        buffer = 20
        y = random.randint(buffer, height - buffer - 1)
        x = random.randint(buffer, width - buffer - 1)
        # Size
        s = random.randint(buffer, height//4)
        return shape, color, (x, y, s)

    def random_image(self, height, width):
        """Creates random specifications of an image with multiple shapes.
        Returns the background color of the image and a list of shape
        specifications that can be used to draw the image.
        """
        # Pick random background color
        bg_color = np.array([random.randint(0, 255) for _ in range(3)])
        # Generate a few random shapes and record their
        # bounding boxes
        shapes = []
        boxes = []
        N = random.randint(1, 4)
        for _ in range(N):
            shape, color, dims = self.random_shape(height, width)
            shapes.append((shape, color, dims))
            x, y, s = dims
            boxes.append([y-s, x-s, y+s, x+s])
        # Apply non-max suppression wit 0.3 threshold to avoid
        # shapes covering each other
        keep_ixs = utils.non_max_suppression(np.array(boxes), np.arange(N), 0.3)
        shapes = [s for i, s in enumerate(shapes) if i in keep_ixs]
        return bg_color, shapes
'''

class MicrocontrollerDataset(utils.Dataset):
    def load_dataset_only_image(self, dataset_dir, augmentations, category_name):
        
        filenames = []
        for i, filename in enumerate(os.listdir(dataset_dir)):
            if '_00.png' in filename:
                filenames.append(filename)
        print(filenames)
        return filenames
        
    def load_dataset(self, dataset_dir, augmentations, category_name):
        
        for i in range(len(category_name)):
            self.add_class('dataset', i, category_name[i])
        
        print('dataset_dir:', dataset_dir.split('/')[-1])
        
        self.dataset_dir = dataset_dir.split('/')[-1]
        
        workbook = Workbook()
        worksheet = workbook.active
        
        rows = 1
        
        worksheet['A1'] = 'index'
        worksheet['B1'] = 'PK_L'
        worksheet['C1'] = 'RGY'
        worksheet['D1'] = 'phase'
        worksheet['E1'] = 'positive_or_negative'
                
        filenames = []
        
        for i, filename in enumerate(os.listdir(dataset_dir)):
            if '_00.png' in filename:
                with open('oral_cancer_datasets/' + dataset_dir.split('/')[-1] + '/lesion_oralCa_20200319~20200924_' + filename.split('/')[-1].split('_')[0].split('.')[0] + '_merge.json', 'r') as json_file:
                    img_anns = json.load(json_file)
                
                worksheet['A{}'.format(rows)] = rows
                worksheet['B{}'.format(rows)] = str(img_anns['filename'].split('/')[-1].split('_')[0] + '_00.png')
                
                if img_anns['category_id'] == 0:
                    worksheet['C{}'.format(rows)] = 'G'
                if img_anns['category_id'] == 1:
                    worksheet['C{}'.format(rows)] = 'Y'
                if img_anns['category_id'] == 2:
                    worksheet['C{}'.format(rows)] = 'R'
    
                worksheet['D{}'.format(rows)] = dataset_dir.split('/')[-1]
                
                if len(img_anns['regions'][0]) == 0:
                    worksheet['E{}'.format(rows)] = 'Negative'
                    rows += 1
                else:
                  worksheet['E{}'.format(rows)] = 'Positive'
                  rows += 1
                  
                  for j in range(len(img_anns['regions'][0])):
                    worksheet['A{}'.format(rows)] = rows
                    worksheet['B{}'.format(rows)] = str(img_anns['filename'].split('/')[-1].split('_')[0] + '_0{}.png'.format(j+1))
                    worksheet['C{}'.format(rows)] = str(img_anns['regions'][0]['region_{}'.format(j+1)][0]['category'])
                    worksheet['D{}'.format(rows)] = dataset_dir.split('/')[-1]
                    worksheet['E{}'.format(rows)] = 'Positive'
                    rows += 1
                    
                self.add_image('dataset', 
                               image_id=i, 
                               path=os.path.join(dataset_dir, filename), 
                               annotation=os.path.join(dataset_dir, 'lesion_oralCa_20200319~20200924_' + filename.split('/')[-1].split('_')[0].split('.')[0] + '_merge.json'))
                filenames.append(filename)
            elif filename.split('_')[-1].split('.')[0] in augmentations and dataset_dir.split('/')[-1] == 'train':
            
                nn = filename.split('_')[-1].split('.')[0]
                
                with open('oral_cancer_datasets/' + dataset_dir.split('/')[-1] + '/lesion_oralCa_20200319~20200924_' + filename.split('/')[-1].split('_')[0].split('.')[0] + '_augmentation_' + nn + '_merge.json', encoding = 'gbk') as json_file:
                    img_anns = json.load(json_file)
                
                worksheet['A{}'.format(rows)] = rows
                worksheet['B{}'.format(rows)] = str(img_anns['filename'].split('/')[-1].split('_')[0] + '_00_' + nn + '.png')
                
                if img_anns['category_id'] == 0:
                    worksheet['C{}'.format(rows)] = 'G'
                if img_anns['category_id'] == 1:
                    worksheet['C{}'.format(rows)] = 'Y'
                if img_anns['category_id'] == 2:
                    worksheet['C{}'.format(rows)] = 'R'
    
                worksheet['D{}'.format(rows)] = dataset_dir.split('/')[-1]
                
                if len(img_anns['regions'][0]) == 0:
                    worksheet['E{}'.format(rows)] = 'Negative'
                    rows += 1
                else:
                  worksheet['E{}'.format(rows)] = 'Positive'
                  rows += 1
                  
                  for j in range(len(img_anns['regions'][0])):
                    worksheet['A{}'.format(rows)] = rows
                    worksheet['B{}'.format(rows)] = str(img_anns['filename'].split('/')[-1].split('_')[0] + '_0{}_{}.png'.format(j+1, nn))
                    worksheet['C{}'.format(rows)] = str(img_anns['regions'][0]['region_{}'.format(j+1)][0]['category'])
                    worksheet['D{}'.format(rows)] = dataset_dir.split('/')[-1]
                    worksheet['E{}'.format(rows)] = 'Positive'
                    rows += 1
            
                self.add_image('dataset', 
                               image_id=i, 
                               path=os.path.join(dataset_dir, filename), 
                               annotation=os.path.join(dataset_dir, 'lesion_oralCa_20200319~20200924_' + filename.split('/')[-1].split('_')[0].split('.')[0] + '_augmentation_' + nn + '_merge.json'))
                filenames.append(filename)

        workbook.save('share_oralCa_ai_cases_' + dataset_dir.split('/')[-1] + '.xlsx')
        
        return filenames
        
    def write_dataset(self, train_dir, val_dir, test_dir):
        
        book_train = openpyxl.load_workbook('share_oralCa_ai_cases_train.xlsx')
        
        sheet_train = book_train.active
        
        booka = Workbook()
        
        sheeta = booka.active
        
        max_row = sheet_train.max_row
        
        num_images_G = 0
        num_images_Y = 0
        num_images_R = 0
        num_lesions_G = 0
        num_lesions_Y = 0
        num_lesions_R = 0
        num_positive_images = 0
        num_negative_images = 0
        num_positive_lesions = 0
        num_negative_lesions = 0
        
        sheeta.cell(row = 1, column = 1).value = 'index'
        sheeta.cell(row = 1, column = 2).value = 'PK_L'
        sheeta.cell(row = 1, column = 3).value = 'RGY'
        sheeta.cell(row = 1, column = 4).value = 'phase'
        sheeta.cell(row = 1, column = 5).value = 'positive_or_negative'

        for i in tqdm(range(2, max_row+1)):
            sheeta.cell(row = i, column = 1).value = i-1
            sheeta.cell(row = i, column = 2).value = sheet_train.cell(row = i, column = 2).value
            sheeta.cell(row = i, column = 3).value = sheet_train.cell(row = i, column = 3).value
            sheeta.cell(row = i, column = 4).value = sheet_train.cell(row = i, column = 4).value
            sheeta.cell(row = i, column = 5).value = sheet_train.cell(row = i, column = 5).value
            
            if str(sheet_train.cell(row = i, column = 2).value).split('_')[-1] == '00.png' or str(sheet_train.cell(row = i, column = 2).value).split('_')[1] == '00':
                
                if str(sheet_train.cell(row = i, column = 3).value) == 'G':
                    num_images_G += 1
                    if str(sheet_train.cell(row = i, column = 5).value) == 'Positive':
                        num_positive_images += 1
                    if str(sheet_train.cell(row = i, column = 5).value) == 'Negative':
                        num_negative_images += 1
                if str(sheet_train.cell(row = i, column = 3).value) == 'Y':
                    num_images_Y += 1
                if str(sheet_train.cell(row = i, column = 3).value) == 'R':
                    num_images_R += 1
            else:
                if str(sheet_train.cell(row = i, column = 3).value) == 'G':
                    num_lesions_G += 1
                    if str(sheet_train.cell(row = i, column = 5).value) == 'Positive':
                        num_positive_lesions += 1
                    if str(sheet_train.cell(row = i, column = 5).value) == 'Negative':
                        num_negative_lesions += 1
                if str(sheet_train.cell(row = i, column = 3).value) == 'Y':
                    num_lesions_Y += 1
                if str(sheet_train.cell(row = i, column = 3).value) == 'R':
                    num_lesions_R += 1       
        
        sheeta.cell(row=1, column=6).value = 'class'
        sheeta.cell(row=1, column=7).value = 'number of images (training phase)'
        sheeta.cell(row=2, column=6).value = 'G'
        sheeta.cell(row=3, column=6).value = 'Y'
        sheeta.cell(row=4, column=6).value = 'R'
        sheeta.cell(row=5, column=6).value = 'Total'
        sheeta.cell(row=2, column=7).value = str(num_images_G) + '(' + str(num_positive_images) + '/' + str(num_negative_images) + ')'
        sheeta.cell(row=3, column=7).value = num_images_Y
        sheeta.cell(row=4, column=7).value = num_images_R
        sheeta.cell(row=5, column=7).value = str(num_images_G + num_images_Y + num_images_R) + '(' + str(num_positive_images + num_images_Y + num_images_R) + '/' + str(num_negative_images) + ')'
        sheeta.cell(row=1, column=8).value = 'number of lesions (training phase)'
        sheeta.cell(row=2, column=8).value = str(num_lesions_G)
        sheeta.cell(row=3, column=8).value = num_lesions_Y
        sheeta.cell(row=4, column=8).value = num_lesions_R
        sheeta.cell(row=5, column=8).value = num_lesions_G + num_lesions_Y + num_lesions_R
        
        book_val = openpyxl.load_workbook('share_oralCa_ai_cases_val.xlsx')
        
        sheet_val = book_val.active
        
        max_row_val = sheet_val.max_row
        
        num_images_G = 0
        num_images_Y = 0
        num_images_R = 0
        num_lesions_G = 0
        num_lesions_Y = 0
        num_lesions_R = 0
        num_positive_images = 0
        num_negative_images = 0
        num_positive_lesions = 0
        num_negative_lesions = 0
        
        for i, j in zip(range(max_row+1, max_row+max_row_val+2), tqdm(range(2, max_row_val+1))):
            sheeta.cell(row = i, column = 1).value = i-1
            sheeta.cell(row = i, column = 2).value = sheet_val.cell(row = j, column = 2).value
            sheeta.cell(row = i, column = 3).value = sheet_val.cell(row = j, column = 3).value
            sheeta.cell(row = i, column = 4).value = sheet_val.cell(row = j, column = 4).value
            sheeta.cell(row = i, column = 5).value = sheet_val.cell(row = j, column = 5).value
            
            if str(sheet_val.cell(row = j, column = 2).value).split('_')[-1] == '00.png':
                if str(sheet_val.cell(row = j, column = 3).value) == 'G':
                    num_images_G += 1
                    if str(sheet_val.cell(row = j, column = 5).value) == 'Positive':
                        num_positive_images += 1
                    if str(sheet_val.cell(row = j, column = 5).value) == 'Negative':
                        num_negative_images += 1
                if str(sheet_val.cell(row = j, column = 3).value) == 'Y':
                    num_images_Y += 1
                if str(sheet_val.cell(row = j, column = 3).value) == 'R':
                    num_images_R += 1
            else:
                if str(sheet_val.cell(row = j, column = 3).value) == 'G':
                    num_lesions_G += 1
                    if str(sheet_val.cell(row = j, column = 5).value) == 'Positive':
                        num_positive_lesions += 1
                    if str(sheet_val.cell(row = j, column = 5).value) == 'Negative':
                        num_negative_lesions += 1
                if str(sheet_val.cell(row = j, column = 3).value) == 'Y':
                    num_lesions_Y += 1
                if str(sheet_val.cell(row = j, column = 3).value) == 'R':
                    num_lesions_R += 1   
        
        sheeta.cell(row=1, column=9).value = 'number of images (validation phase)'
        sheeta.cell(row=2, column=9).value = str(num_images_G) + '(' + str(num_positive_images) + '/' + str(num_negative_images) + ')'
        sheeta.cell(row=3, column=9).value = num_images_Y
        sheeta.cell(row=4, column=9).value = num_images_R
        sheeta.cell(row=5, column=9).value = str(num_images_G + num_images_Y + num_images_R) + '(' + str(num_positive_images + num_images_Y + num_images_R) + '/' + str(num_negative_images) + ')'
        sheeta.cell(row=1, column=10).value = 'number of lesions (validation phase)'
        sheeta.cell(row=2, column=10).value = str(num_lesions_G)
        sheeta.cell(row=3, column=10).value = num_lesions_Y
        sheeta.cell(row=4, column=10).value = num_lesions_R
        sheeta.cell(row=5, column=10).value = num_lesions_G + num_lesions_Y + num_lesions_R
        
        book_test = openpyxl.load_workbook('share_oralCa_ai_cases_test.xlsx')
        
        sheet_test = book_test.active
        
        max_row_test = sheet_test.max_row
        
        num_images_G = 0
        num_images_Y = 0
        num_images_R = 0
        num_lesions_G = 0
        num_lesions_Y = 0
        num_lesions_R = 0
        num_positive_images = 0
        num_negative_images = 0
        num_positive_lesions = 0
        num_negative_lesions = 0
        
        for i, j in zip(range(max_row+max_row_val+2, max_row+max_row_val+max_row_test+3), tqdm(range(2, max_row_test+1))):
            sheeta.cell(row = i, column = 1).value = i-1
            sheeta.cell(row = i, column = 2).value = sheet_test.cell(row = j, column = 2).value
            sheeta.cell(row = i, column = 3).value = sheet_test.cell(row = j, column = 3).value
            sheeta.cell(row = i, column = 4).value = sheet_test.cell(row = j, column = 4).value
            sheeta.cell(row = i, column = 5).value = sheet_test.cell(row = j, column = 5).value
            
            if str(sheet_test.cell(row = j, column = 2).value).split('_')[-1] == '00.png':
                if str(sheet_test.cell(row = j, column = 3).value) == 'G':
                    num_images_G += 1
                    if str(sheet_test.cell(row = j, column = 5).value) == 'Positive':
                        num_positive_images += 1
                    if str(sheet_test.cell(row = j, column = 5).value) == 'Negative':
                        num_negative_images += 1
                if str(sheet_test.cell(row = j, column = 3).value) == 'Y':
                    num_images_Y += 1
                if str(sheet_test.cell(row = j, column = 3).value) == 'R':
                    num_images_R += 1
            else:
                if str(sheet_test.cell(row = j, column = 3).value) == 'G':
                    num_lesions_G += 1
                    if str(sheet_test.cell(row = j, column = 5).value) == 'Positive':
                        num_positive_lesions += 1
                    if str(sheet_test.cell(row = j, column = 5).value) == 'Negative':
                        num_negative_lesions += 1
                if str(sheet_test.cell(row = j, column = 3).value) == 'Y':
                    num_lesions_Y += 1
                if str(sheet_test.cell(row = j, column = 3).value) == 'R':
                    num_lesions_R += 1   
        
        sheeta.cell(row=1, column=11).value = 'number of images (testing phase)'
        sheeta.cell(row=2, column=11).value = str(num_images_G) + '(' + str(num_positive_images) + '/' + str(num_negative_images) + ')'
        sheeta.cell(row=3, column=11).value = num_images_Y
        sheeta.cell(row=4, column=11).value = num_images_R
        sheeta.cell(row=5, column=11).value = str(num_images_G + num_images_Y + num_images_R) + '(' + str(num_positive_images + num_images_Y + num_images_R) + '/' + str(num_negative_images) + ')'
        sheeta.cell(row=1, column=12).value = 'number of lesions (testing phase)'
        sheeta.cell(row=2, column=12).value = str(num_lesions_G)
        sheeta.cell(row=3, column=12).value = num_lesions_Y
        sheeta.cell(row=4, column=12).value = num_lesions_R
        sheeta.cell(row=5, column=12).value = num_lesions_G + num_lesions_Y + num_lesions_R
        
        booka.save('share_oralCa_ai_cases_total_analysis.xlsx')
    
    # extract bounding boxes from an annotation file
    def extract_boxes(self, filename):
        
        # load and parse the file
        with open('oral_cancer_datasets/' + self.dataset_dir + '/lesion_oralCa_20200319~20200924_' + filename.split('/')[-1].split('_')[0].split('.')[0] + '_merge.json', 'r') as json_file:
            annotation = json.load(json_file)
        
        boxes = []
        classes = []
        
        for j in range(1, 1+len(annotation['regions'][0])):
            if annotation['regions'][0]['region_{}'.format(j)][0]['category'] == 'G':
                classes.append('1')
            if annotation['regions'][0]['region_{}'.format(j)][0]['category'] == 'Y':
                classes.append('2')
            if annotation['regions'][0]['region_{}'.format(j)][0]['category'] == 'R':
                classes.append('3')
                    
            boxes.append(annotation['regions'][0]['region_{}'.format(j)][0]['bbox'])
        
        # extract image dimensions
        width = 384
        height = 512
        return boxes, classes, width, height
    
    def extract_masks(self, filename, image_id):
        
        with open(filename, 'r') as json_file:
            img_anns = json.load(json_file)
            
        masks = np.zeros([384, 512, len(img_anns['regions'][0])], dtype='uint8')
        classes = []
        
        image = cv2.imread('oral_cancer_datasets/' + self.dataset_dir + '/' + img_anns['filename'].split('/')[-1])
    
        for j in range(len(img_anns['regions'][0])):
            mask = np.zeros((384, 512), np.uint8)
            polygons = []
            
            for i in range(len(img_anns['regions'][0]['region_{}'.format(j+1)][0]['all_points_x'])):
                polygons.append([img_anns['regions'][0]['region_{}'.format(j+1)][0]['all_points_y'][i], img_anns['regions'][0]['region_{}'.format(j+1)][0]['all_points_x'][i]])
            
            cv2.fillPoly(mask, [np.array(polygons, dtype=np.int32)], 1)
            masks[:, :, j] = mask
            
            if img_anns['regions'][0]['region_{}'.format(j+1)][0]['category'] == 'G':
                classes.append('1')
            if img_anns['regions'][0]['region_{}'.format(j+1)][0]['category'] == 'Y':
                classes.append('2')
            if img_anns['regions'][0]['region_{}'.format(j+1)][0]['category'] == 'R':
                classes.append('3')
            
        return masks, classes
        
    # load the masks for an image
    def load_mask(self, image_id):
        
        info = self.image_info[image_id]
        # define box file location
        path = info['annotation']
        # load XML
        # boxes, classes, w, h = self.extract_boxes(path)
        masks, classes = self.extract_masks(path, image_id)
        
        return masks, np.asarray(classes, dtype='int32')
    
    def image_reference(self, image_id):
        info = self.image_info[image_id]
        return info['path']

args = opt()
config = ShapesConfig()
config.BACKBONE = args.backbone

if config.BACKBONE == 'resnet50' or config.BACKBONE == 'resnet101' or config.BACKBONE == 'resnet152':
    config.MINI_MASK_SHAPE = (28, 28)  # (height, width) of the mini-mask
elif config.BACKBONE == 'densenet' or config.BACKBONE == 'mobilenetv1' or config.BACKBONE == 'mobilenetv2':
    config.MINI_MASK_SHAPE = (56, 56)
if config.BACKBONE == 'resnet50' or config.BACKBONE == 'resnet101' or config.BACKBONE == 'mobilenetv1' or config.BACKBONE == 'mobilenetv2':
    config.LOSS_WEIGHTS = {
        "rpn_class_loss": 1.,
        "rpn_bbox_loss": 1.,
        "mrcnn_class_loss": 1.,
        "mrcnn_bbox_loss": 1.,
        "mrcnn_mask_loss": 1.
    }
elif config.BACKBONE == 'densenet':
    config.LOSS_WEIGHTS = {
        "rpn_class_loss": 1.,
        "rpn_bbox_loss": 1.,
        "mrcnn_class_loss": 1.,
        "mrcnn_bbox_loss": 1.,
        "keypoint_mrcnn_mask_loss": 1.,
        "mrcnn_mask_loss": 1.
    }
    
config.display()

'''
# Training dataset
dataset_train = ShapesDataset()
dataset_train.load_shapes(500, config.IMAGE_SHAPE[0], config.IMAGE_SHAPE[1])
dataset_train.prepare()

# Validation dataset
dataset_val = ShapesDataset()
dataset_val.load_shapes(50, config.IMAGE_SHAPE[0], config.IMAGE_SHAPE[1])
dataset_val.prepare()

# Load and display random samples
image_ids = np.random.choice(dataset_train.image_ids, 4)
'''

'''
for image_id in image_ids:
    image = dataset_train.load_image(image_id)
    mask, class_ids = dataset_train.load_mask(image_id)
    visualize.display_top_masks(image, mask, class_ids, dataset_train.class_names)
'''

# adjusting
if config.BACKBONE == 'densenet':
    from mrcnn.model_densenet import log
    import mrcnn.model_densenet as modellib
    
if config.BACKBONE == 'resnet50' or config.BACKBONE == 'resnet101' or config.BACKBONE == 'resnet152':
    #from mrcnn_v2.model_resnet import log
    #import mrcnn_v2.model_resnet as modellib
    if args.mask_mode == 'Cascade_MaskRCNN':
        from mrcnn.model_resnet_cascade import log
        import mrcnn.model_resnet_cascade as modellib
    else:
        from mrcnn.model_resnet import log
        import mrcnn.model_resnet as modellib

# adjusting
if config.BACKBONE == 'mobilenetv1' or config.BACKBONE == 'mobilenetv2':
    from mrcnn.model_mobilenet import log
    import mrcnn.model_mobilenet as modellib

# adjusting
if config.BACKBONE == 'xception':
    from mrcnn.model_xception import log
    import mrcnn.model_xception as modellib
    
# Create training and validation set
# train set
train_dir = 'oral_cancer_datasets/train'
dataset_train = MicrocontrollerDataset()
file_names_train = dataset_train.load_dataset(train_dir, config.augmentation_types, config.category_name)
dataset_train.prepare()
print('Train: %d' % len(dataset_train.image_ids))

# val set
val_dir = 'oral_cancer_datasets/val'
dataset_val = MicrocontrollerDataset()
file_names_val = dataset_val.load_dataset(val_dir, config.augmentation_types, config.category_name)
dataset_val.prepare()
print('Val: %d' % len(dataset_val.image_ids))

# test set
test_dir = 'oral_cancer_datasets/test_20220210'
dataset_test = MicrocontrollerDataset()
file_names_test = dataset_test.load_dataset(test_dir, config.augmentation_types, config.category_name)
dataset_test.prepare()
print('Test: %d' % len(dataset_test.image_ids))

# autolabel set
autolabel_dir = 'oral_cancer_datasets/autolabel_2'
dataset_autolabel = MicrocontrollerDataset()
file_names_autolabel = dataset_autolabel.load_dataset_only_image(autolabel_dir, config.augmentation_types, config.category_name)
dataset_autolabel.prepare()
print('autolabel: %d' % len(dataset_autolabel.image_ids))

# statistical analysis of total dataset
MicrocontrollerDataset().write_dataset(train_dir, val_dir, test_dir)

'''
# Which weights to start with?
init_with = "coco"  # imagenet, coco, or last

if init_with == "imagenet":
    model.load_weights(model.get_imagenet_weights(), by_name=True)
elif init_with == "coco":
    # Load weights trained on MS COCO, but skip layers that
    # are different due to the different number of classes
    # See README for instructions to download the COCO weights
    model.load_weights(COCO_MODEL_PATH, by_name=True,
                       exclude=["mrcnn_class_logits", "mrcnn_bbox_fc", 
                                "mrcnn_bbox", "mrcnn_mask"])
elif init_with == "last":
    # Load the last model you trained and continue training
    model.load_weights(model.find_last(), by_name=True)


# Train the head branches
# Passing layers="heads" freezes all layers except the head
# layers. You can also pass a regular expression to select
# which layers to train by name pattern.
model.train(dataset_train, dataset_val, 
            learning_rate=config.LEARNING_RATE, 
            epochs=1, 
            layers='heads')
            
# Fine tune all layers
# Passing layers="all" trains all layers. You can also 
# pass a regular expression to select which layers to
# train by name pattern.
model.train(dataset_train, dataset_val, 
            learning_rate=config.LEARNING_RATE / 10,
            epochs=2, 
            layers="all")
'''

args = opt()

os.environ["CUDA_VISIBLE_DEVICES"] = args.device

if args.training_or_inference_mode == 'training':
    if args.mask_mode == 'MaskRCNN':
        # Create model in training mode
        config.Name = 'microcontroller_maskrcnn_detection'
        model = modellib.MaskRCNN(mode="training", config=config, model_dir=MODEL_DIR)
    elif args.mask_mode == 'Cascade_MaskRCNN':
        config.Name = 'microcontroller_cascade_maskrcnn_detection'
        model = modellib.Cascade_MaskRCNN(mode="training", config=config, model_dir=MODEL_DIR)
    
    if args.model_path:
        # Load the last model you trained and continue training
        epoch=int(args.model_path.split('_')[-1].split('.')[0])
        print("Loading weights from ", args.model_path)
        model.load_weights(args.model_path, by_name=True)
    else:
        epoch=0
        model.load_weights(model.get_imagenet_weights(), by_name=True)
    
    print('Prepare for training... (from epoch=', epoch,')')
    # model = ParallelModel(model, 2)
    
    #model.train(dataset_train, dataset_val, 
    #        learning_rate=config.LEARNING_RATE / 10,
    #        init_epoch=epoch,
    #        epochs=200, 
    #        layers="all")
            
    model.train(dataset_train, dataset_val, 
            learning_rate=config.LEARNING_RATE / 10,
            init_epoch=epoch,
            epochs=1000, 
            layers="all",
            augmentation = imgaug.augmenters.Sequential([ 
                imgaug.augmenters.Fliplr(1), 
                imgaug.augmenters.Flipud(1), 
                imgaug.augmenters.Affine(rotate=(-45, 45)), 
                imgaug.augmenters.Affine(rotate=(-90, 90)), 
                imgaug.augmenters.Affine(scale=(0.5, 1.5))]))

elif args.training_or_inference_mode == 'inference':
    
    class InferenceConfig(Config):
        GPU_COUNT = 1
        IMAGES_PER_GPU = 1
        BACKBONE = args.backbone

    inference_config = InferenceConfig()
    
    if args.mask_mode == 'MaskRCNN':
        model = modellib.MaskRCNN(mode="inference", config=config, model_dir=MODEL_DIR)
    elif args.mask_mode == 'Cascade_MaskRCNN':
        model = modellib.Cascade_MaskRCNN(mode="inference", config=config, model_dir=MODEL_DIR)                      
    if not os.path.exists('results_' + args.mask_mode + '_' + args.backbone):
        os.makedirs('results_' + args.mask_mode + '_' + args.backbone)
        
    print("Loading weights from ", args.model_path)
    model.load_weights(args.model_path, by_name=True)

    # Compute VOC-Style mAP @ IoU=0.5
    # Running on : images. Increase for better accuracy.
    
    image_ids = dataset_test.image_ids
    # print('image_ids:', image_ids)
    
    pbar = tqdm(range(len(image_ids)))
    colors = config.colors
    
    #APs = []
    #ground-truth and predictions lists
    gt_tot_per_image = np.array([])
    pred_tot_per_image = np.array([])
    gt_tot = np.array([])
    pred_tot = np.array([])
    #mAP list
    mAP_ = []
    mprecision_ = []
    mrecall_ = []
    mdice = []

    count1 = 0
    
    class_names = []
    class_names_index = []
    
    for image_id, i in zip(image_ids, pbar):
        # Load image and ground truth data
        
        image, image_meta, gt_class_id, gt_bbox, gt_mask =\
                modellib.load_image_gt(dataset_test, Config, image_id, use_mini_mask=False)

        molded_image = np.expand_dims(image, 0)
        
        results = model.detect(molded_image, verbose=0)
        r = results[0]
        print('Processing file results_' + args.mask_mode + '/' + str(file_names_test[i]))
        
        if count1 == 0:
            save_box, save_class, save_mask = gt_bbox, gt_class_id, gt_mask
            save_roi, save_id, save_score, save_m = r["rois"], r["class_ids"], r["scores"], r['masks']
        else:
            save_roi = np.concatenate((save_roi, r["rois"]), axis=0)
            save_id = np.concatenate((save_id, r["class_ids"]), axis=0)
            save_score = np.concatenate((save_score, r["scores"]), axis=0)
            save_m = np.concatenate((save_m, r['masks']), axis=2)
            save_box = np.concatenate((save_box, gt_bbox), axis=0)
            save_class = np.concatenate((save_class, gt_class_id), axis=0)
            save_mask = np.concatenate((save_mask, gt_mask), axis=2)
        
        count1 += 1
        
        #compute gt_tot and pred_tot
        gt, pred = utils.gt_pred_lists(gt_class_id, gt_bbox, r['class_ids'], r['rois'])
        gt_tot = np.append(gt_tot, gt)
        pred_tot = np.append(pred_tot, pred)
        
        if len(gt_class_id) == 0 and len(r['class_ids']) > 0:
            gt_tot_per_image = np.append(gt_tot_per_image, np.array(0))
            pred_tot_per_image = np.append(pred_tot_per_image, np.max(r['class_ids']))
            cls = np.max(r['class_ids'])
            if dataset_test.class_names[int(cls)] not in class_names:
                class_names.append(dataset_test.class_names[int(cls)])
                class_names_index.append(int(cls))
                
        elif len(gt_class_id) > 0 and len(r['class_ids']) == 0:
            gt_tot_per_image = np.append(gt_tot_per_image, np.max(gt_class_id))
            pred_tot_per_image = np.append(pred_tot_per_image, np.array(0))
            cls = np.max(gt_class_id)
            if dataset_test.class_names[int(cls)] not in class_names:
                class_names.append(dataset_test.class_names[int(cls)])
                class_names_index.append(int(cls))
                
        elif len(r['class_ids']) == 0 and len(gt_class_id) == 0:
            gt_tot_per_image = np.append(gt_tot_per_image, np.array(0))
            pred_tot_per_image = np.append(pred_tot_per_image, np.array(0))
            if dataset_test.class_names[0] not in class_names:
                class_names.append(dataset_test.class_names[0])
                class_names_index.append(0)
                
        else:
            gt_tot_per_image = np.append(gt_tot_per_image, np.max(gt_class_id))
            pred_tot_per_image = np.append(pred_tot_per_image, np.max(r['class_ids']))
            cls = np.max(gt_class_id)
            if dataset_test.class_names[int(cls)] not in class_names:
                class_names.append(dataset_test.class_names[int(cls)])
                class_names_index.append(int(cls))
            
        color_list = []
        cls_index = []
        
        for m, cls_name in enumerate(r['class_ids']):
            if cls_name == 3:
                color_list.append(colors[2])
            if cls_name == 2:
                color_list.append(colors[1])
            if cls_name == 1:
                color_list.append(colors[0])
            cls_index.append(cls_name)
        
        dice = []
        for j in range(len(gt_class_id)):
            gt = gt_class_id[j].astype(np.int) * np.array(gt_mask.astype(np.int)[:,:,j])
            for k in range(len(cls_index)):
                seg = cls_index[k].astype(np.int) * r["masks"].astype(np.int)[:,:,k]
                score = np.sum(seg[gt==gt_class_id[j].astype(np.int)])*2.0 / (np.sum(seg) + np.sum(gt))
    
                if score <= 1:
                    dice.append(score)
        
        if np.mean(dice) <= 1:
            # print('Dice similarity score is {}'.format(np.mean(dice)))
            mdice.append(np.mean(dice))
        image_display = visualize.display_instances(image, r['rois'], r['masks'], r['class_ids'], 
                            dataset_test.class_names, r['scores'], ax=get_ax(), colors=color_list)
                            
        # visualize.plot_precision_recall(AP, precisions, recalls)
        plt.imshow(image_display)
        #plt.show()
        plt.savefig('results_' + args.mask_mode + '_' + args.backbone + '/' + str(file_names_test[i]), bbox_inches='tight')
        plt.close()
        
    mdice = np.mean(mdice)
    #print('Mean dice coefficient:', mdice)
    
    class_names_index = sorted(class_names_index)
    
    class_names = []
    
    for i in class_names_index:
        class_names.append(dataset_test.class_names[int(i)])
        
    utils.plot_confusion_matrix_from_data_per_image(gt_tot_per_image, pred_tot_per_image, args.backbone, args.model_path.split('/')[-1].split('.')[0], class_names, fz=18, figsize=(20,20), lw=0.5)
    
    fig, ax = plt.subplots(1, 1, figsize=(9, 6), tight_layout=True)
    
    colorlist = config.colorlist
    
    F1_scores = []
    
    for class_id in range(1, len(colorlist)):
        mAP, precisions, recalls, _ =\
                utils.compute_ap_pre_class(save_box, save_class, save_mask,
                         save_roi, save_id, save_score, save_m, class_id)
        mAR, _ = utils.compute_recall(save_roi, gt_bbox, iou=0.5) 
        
        #visualize.plot_precision_recall(mmAP, mprecisions, mrecalls, args.backbone)
        
        F1_scores = (2* (np.mean(precisions) * np.mean(recalls)))/(np.mean(precisions) + np.mean(recalls))
        
        ax.plot(recalls, precisions, color = colorlist[class_id], label=dataset_test.class_names[class_id] + ' mAP=%.2f, F1 score=%.2f' % (mAP, F1_scores))
        
    ax.set_xlabel('Recall (Sensitivity, True positive rate)')
    ax.set_ylabel('Precision (Positive predicted values)')
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    
    ax.legend(loc="upper right", prop={'size': 12})
            
    # Compute AP
    mmAP, mprecisions, mrecalls, overlaps =\
                utils.compute_ap(save_box, save_class, save_mask,
                         save_roi, save_id, save_score, save_m)
    mmAR, _ = utils.compute_recall(save_roi, gt_bbox, iou=0.5) 
    F1_scores = (2* (np.mean(mprecisions) * np.mean(mrecalls)))/(np.mean(mprecisions) + np.mean(mrecalls))
    ax.plot(mrecalls, mprecisions, color = 'Blue', label='Total mAP=%.2f, F1 score=%.2f, Mean dice similarity score=%.2f' % (mmAP, F1_scores, mdice))
    
    #visualize.plot_precision_recall(mmAP, mprecisions, mrecalls, args.backbone)
    plt.legend()
    plt.savefig('results_pr_curves_{}_{}.png'.format(args.backbone, args.model_path.split('/')[-1].split('.')[0]) , bbox_inches='tight')
    
elif args.training_or_inference_mode == 'autolabel':
    
    class InferenceConfig(Config):
        GPU_COUNT = 1
        IMAGES_PER_GPU = 1
        BACKBONE = args.backbone

    inference_config = InferenceConfig()
    
    if args.mask_mode == 'MaskRCNN':
        model = modellib.MaskRCNN(mode="inference", config=config, model_dir=MODEL_DIR)
    elif args.mask_mode == 'CasCadeMaskRCNN':
        model = modellib.CasCade_MaskRCNN(mode="inference", config=config, model_dir=MODEL_DIR)                      
    if not os.path.exists('results_' + args.mask_mode + '_' + args.backbone):
        os.makedirs('results_' + args.mask_mode + '_' + args.backbone)
        
    print("Loading weights from ", args.model_path)
    model.load_weights(args.model_path, by_name=True)

    # Compute VOC-Style mAP @ IoU=0.5
    pbar = tqdm(range(len(file_names_autolabel)))
    colors = config.colors
    
    gt_tot_per_image = np.array([])
    pred_tot_per_image = np.array([])
    gt_tot = np.array([])
    pred_tot = np.array([])
    APs = []
    ARs = []
    F1_scores = []
    count1 = 0
    
    for image_id, i in zip(file_names_autolabel, pbar):
        # Load image and ground truth data
        #print('image_id:', image_id)
        image = np.array(cv2.cvtColor(cv2.imread('oral_cancer_datasets/autolabel/' + image_id), cv2.COLOR_BGR2RGB))
        #molded_images = np.expand_dims(modellib.mold_image(image, inference_config), 0)
        #print('shape:', image.shape)
        molded_image = np.expand_dims(image, 0)
        # Run object detection
        results = model.detect(molded_image, verbose=0)
        r = results[0]
        
        '''
        visualize.display_differences(image,
                        gt_bbox, gt_class_id, gt_mask,
                        r["rois"], r["class_ids"], r["scores"], r['masks'],
                        dataset_test.class_names, title="", ax=None,
                        show_mask=True, show_box=True,
                        iou_threshold=0.5, score_threshold=0.5)
        '''
        
        color_list = []
        for cls_name in r['class_ids']:
            if cls_name == 3:
                color_list.append(colors[2])
            if cls_name == 2:
                color_list.append(colors[1])
            if cls_name == 1:
                color_list.append(colors[0])
        
        #print('dataset_test.class_names:', dataset_test.class_names)
        
        image_display = visualize.display_instances_auto_label(image, r['rois'], r['masks'], r['class_ids'], 
                            ['B', 'G', 'Y', 'R'], r['scores'], ax=get_ax(), colors=color_list)
        # visualize.plot_precision_recall(AP, precisions, recalls)
        plt.imshow(image_display)
        plt.savefig('results_' + args.mask_mode + '_' + args.backbone + '/' + str(file_names_autolabel[i].split('.')[0] + args.mask_mode + '_' + args.backbone + '.png'), bbox_inches='tight')
        plt.close()
        
        fig, ax = plt.subplots(1,2)
        ax[0].imshow(image)
        ax[1].imshow(image_display)
        plt.title(str(file_names_autolabel[i]))
        plt.axis('off')
        plt.savefig('results_' + args.mask_mode + '_' + args.backbone + '/groundtruth_and_inference_' + str(file_names_autolabel[i]), bbox_inches='tight')
        plt.close()
    
'''
class InferenceConfig(ShapesConfig):
    GPU_COUNT = 1
    IMAGES_PER_GPU = 1

inference_config = InferenceConfig()

# Recreate the model in inference mode
model = modellib.MaskRCNN(mode="inference", 
                          config=inference_config,
                          model_dir=MODEL_DIR)

# Get path to saved weights
# Either set a specific path or find last trained weights
# model_path = os.path.join(ROOT_DIR, ".h5 file name here")
model_path = model.find_last()

# Load trained weights
print("Loading weights from ", model_path)
model.load_weights(model_path, by_name=True)

# Test on a random image
image_id = random.choice(dataset_val.image_ids)
original_image, image_meta, gt_class_id, gt_bbox, gt_mask =\
    modellib.load_image_gt(dataset_val, inference_config, 
                           image_id, use_mini_mask=False)

log("original_image", original_image)
log("image_meta", image_meta)
log("gt_class_id", gt_class_id)
log("gt_bbox", gt_bbox)
log("gt_mask", gt_mask)

visualize.display_instances(original_image, gt_bbox, gt_mask, gt_class_id, 
                            dataset_train.class_names, figsize=(8, 8))

results = model.detect([original_image], verbose=1)

r = results[0]
visualize.display_instances(original_image, r['rois'], r['masks'], r['class_ids'], 
                            dataset_val.class_names, r['scores'], ax=get_ax())

# Compute VOC-Style mAP @ IoU=0.5
# Running on 10 images. Increase for better accuracy.
image_ids = np.random.choice(dataset_val.image_ids, 10)
APs = []
for image_id in image_ids:
    # Load image and ground truth data
    image, image_meta, gt_class_id, gt_bbox, gt_mask =\
        modellib.load_image_gt(dataset_val, inference_config,
                               image_id, use_mini_mask=False)
    molded_images = np.expand_dims(modellib.mold_image(image, inference_config), 0)
    # Run object detection
    results = model.detect([image], verbose=0)
    r = results[0]
    # Compute AP
    AP, precisions, recalls, overlaps =\
        utils.compute_ap(gt_bbox, gt_class_id, gt_mask,
                         r["rois"], r["class_ids"], r["scores"], r['masks'])
    APs.append(AP)
    
print("mAP: ", np.mean(APs))
'''